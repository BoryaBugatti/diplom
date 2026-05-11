# routers/admin_router.py
import json
import csv
import io
import logging
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from pydantic import BaseModel

from config.dbconfig import SessionLocal
from Models.models import User, AnalysisResult, Report
from auth.dependencies import get_current_admin_user


def get_db():
    db = SessionLocal()
    try:
        yield db
        db.commit()
    except:
        db.rollback()
        raise
    finally:
        db.close()

router = APIRouter(prefix="/admin", tags=["Admin"])
logger = logging.getLogger(__name__)

# ---------- Pydantic схемы ----------
class AnalysisResponse(BaseModel):
    id: int
    user_id: int
    user_name: str
    file_name: str
    tender_name: Optional[str]
    tender_description: Optional[str]
    all_requirements: Optional[str]
    key_requirements: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True

class UserResponse(BaseModel):
    user_id: int
    user_name: str
    user_email: str
    user_role: str
    created_at: datetime

class RoleUpdate(BaseModel):
    role: str

class ReportGenerateRequest(BaseModel):
    analysis_id: int
    format: str  # 'pdf', 'csv', 'xlsx', 'json'

# ---------- Вспомогательная функция для безопасного парсинга требований ----------
def safe_parse_requirements(value: Optional[str]) -> List[str]:
    """Преобразует JSON-строку в список строк. При ошибке возвращает пустой список."""
    if not value:
        return []
    try:
        # Пробуем распарсить как JSON
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return parsed
        else:
            return [str(parsed)]
    except json.JSONDecodeError:
        # Если не JSON, возможно это уже строка с пробелами, пытаемся убрать лишние пробелы
        # Это костыль для испорченных данных
        if all(c.isalpha() or c in '.,!?;:' for c in value.replace(' ', '')):
            # Убираем пробелы между символами
            cleaned = value.replace(' ', '')
            return [cleaned]
        return [value]

# ---------- Эндпоинты ----------
@router.get("/analyses", response_model=List[AnalysisResponse])
async def get_all_analyses(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """Получить все анализы всех пользователей."""
    stmt = select(
        AnalysisResult,
        User.user_name
    ).join(User, AnalysisResult.user_id == User.user_id).order_by(AnalysisResult.created_at.desc())
    results = db.execute(stmt).all()
    
    analyses = []
    for analysis, user_name in results:
        analyses.append(AnalysisResponse(
            id=analysis.id,
            user_id=analysis.user_id,
            user_name=user_name,
            file_name=analysis.file_name,
            tender_name=analysis.tender_name,
            tender_description=analysis.tender_description,
            all_requirements=analysis.all_requirements,
            key_requirements=analysis.key_requirements,
            created_at=analysis.created_at
        ))
    return analyses

@router.get("/users", response_model=List[UserResponse])
async def get_all_users(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    users = db.execute(select(User).order_by(User.created_at.desc())).scalars().all()
    return users

@router.put("/users/{user_id}/role")
async def update_user_role(
    user_id: int,
    role_data: RoleUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    if user_id == admin.user_id:
        raise HTTPException(status_code=400, detail="Нельзя изменить собственную роль")
    
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    if role_data.role not in {"user", "admin"}:
        raise HTTPException(status_code=400, detail="Недопустимая роль. Допустимые: user, admin")
    
    old_role = user.user_role
    user.user_role = role_data.role
    db.add(user)
    # commit выполнится автоматически в get_db
    
    return {"message": f"Роль пользователя {user.user_name} изменена с '{old_role}' на '{user.user_role}'"}

@router.post("/reports/generate")
async def generate_report(
    req: ReportGenerateRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """Генерирует отчёт по анализу в указанном формате."""
    # 1. Получаем анализ
    analysis = db.get(AnalysisResult, req.analysis_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Анализ не найден")
    
    # 2. Получаем пользователя
    user = db.get(User, analysis.user_id)
    user_name = user.user_name if user else "Неизвестный"
    
    # 3. Безопасно парсим требования
    all_reqs = safe_parse_requirements(analysis.all_requirements)
    key_reqs = safe_parse_requirements(analysis.key_requirements)
    
    # 4. Данные для отчёта
    data = {
        "id": analysis.id,
        "user_name": user_name,
        "file_name": analysis.file_name,
        "tender_name": analysis.tender_name or "",
        "tender_description": analysis.tender_description or "",
        "created_at": analysis.created_at.isoformat(),
        "all_requirements": all_reqs,
        "key_requirements": key_reqs
    }
    
    format_lower = req.format.lower()
    
    try:
        if format_lower == "json":
            content = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
            media_type = "application/json"
            filename = f"report_{analysis.id}.json"
        
        elif format_lower == "csv":
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(["ID", "Пользователь", "Файл", "Название тендера", "Описание", "Дата", "Все требования", "Ключевые требования"])
            writer.writerow([
                data["id"], data["user_name"], data["file_name"], data["tender_name"],
                data["tender_description"], data["created_at"],
                "; ".join(data["all_requirements"]), "; ".join(data["key_requirements"])
            ])
            content = output.getvalue().encode("utf-8")
            media_type = "text/csv"
            filename = f"report_{analysis.id}.csv"
        
        elif format_lower == "xlsx":
            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(status_code=500, detail="Библиотека openpyxl не установлена. Установите: pip install openpyxl")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Анализ тендера"
            ws.append(["ID", data["id"]])
            ws.append(["Пользователь", data["user_name"]])
            ws.append(["Файл", data["file_name"]])
            ws.append(["Название тендера", data["tender_name"]])
            ws.append(["Описание", data["tender_description"]])
            ws.append(["Дата", data["created_at"]])
            ws.append(["Все требования"] + ([""] * (len(data["all_requirements"]) - 1) if len(data["all_requirements"]) > 1 else []))
            for idx, req in enumerate(data["all_requirements"], start=1):
                ws.cell(row=7, column=idx+1, value=req)
            ws.append(["Ключевые требования"] + ([""] * (len(data["key_requirements"]) - 1) if len(data["key_requirements"]) > 1 else []))
            for idx, req in enumerate(data["key_requirements"], start=1):
                ws.cell(row=8, column=idx+1, value=req)
            
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"report_{analysis.id}.xlsx"
        
        elif format_lower == "pdf":
            try:
                from reportlab.pdfgen import canvas
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.units import mm
            except ImportError:
                raise HTTPException(status_code=500, detail="Библиотека reportlab не установлена. Установите: pip install reportlab")
            
            buffer = io.BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 20*mm
            
            c.setFont("Helvetica-Bold", 14)
            c.drawString(20*mm, y, f"Отчёт по анализу #{data['id']}")
            y -= 10*mm
            
            c.setFont("Helvetica", 10)
            c.drawString(20*mm, y, f"Пользователь: {data['user_name']}")
            y -= 6*mm
            c.drawString(20*mm, y, f"Файл: {data['file_name']}")
            y -= 6*mm
            c.drawString(20*mm, y, f"Название тендера: {data['tender_name']}")
            y -= 6*mm
            
            # Описание (обрезаем при необходимости)
            desc = data['tender_description']
            if len(desc) > 500:
                desc = desc[:500] + "..."
            text = c.beginText(20*mm, y)
            text.setFont("Helvetica", 9)
            text.textLines(f"Описание:\n{desc}")
            c.drawText(text)
            y -= 40*mm
            
            # Требования (только первые 10, чтобы не выходить за страницу)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(20*mm, y, "Все требования:")
            y -= 6*mm
            c.setFont("Helvetica", 8)
            for i, req in enumerate(data["all_requirements"][:10]):
                if y < 30*mm:
                    c.showPage()
                    y = height - 20*mm
                c.drawString(25*mm, y, f"• {req[:80]}")
                y -= 5*mm
            
            c.save()
            content = buffer.getvalue()
            media_type = "application/pdf"
            filename = f"report_{analysis.id}.pdf"
        
        else:
            raise HTTPException(status_code=400, detail=f"Неподдерживаемый формат: {format_lower}. Доступны: json, csv, xlsx, pdf")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка генерации отчёта")
        raise HTTPException(status_code=500, detail=f"Внутренняя ошибка: {str(e)}")
    
    # 5. Сохраняем информацию о генерации отчёта
    new_report = Report(
        admin_id=admin.user_id,
        analysis_id=analysis.id,
        report_format=format_lower,
        file_path=None
    )
    db.add(new_report)
    # commit автоматический
    
    # 6. Возвращаем файл
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/reports/count")
async def get_reports_count(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    count = db.execute(select(func.count(Report.report_id))).scalar_one()
    return {"total_reports": count}

@router.get("/statistics")
async def get_admin_statistics(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    total_analyses = db.execute(select(func.count(AnalysisResult.id))).scalar_one()
    total_users = db.execute(select(func.count(User.user_id))).scalar_one()
    total_reports = db.execute(select(func.count(Report.report_id))).scalar_one()
    return {
        "total_analyses": total_analyses,
        "total_users": total_users,
        "total_reports": total_reports
    }